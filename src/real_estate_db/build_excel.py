from __future__ import annotations

import argparse
import csv
import hashlib
from collections import Counter
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .schema import REGION_ORDER, REQUIRED_COLUMNS, URL_COLUMNS
from .validate import validate_file

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_INPUT = ROOT / "data" / "real_estate_brokers.csv"
DEFAULT_OUTPUT_DIR = ROOT / "database"

HEADER_FILL = PatternFill("solid", fgColor="0B4F8A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
LINK_FONT = Font(color="0563C1", underline="single")
THIN_GREY = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN_GREY, right=THIN_GREY, top=THIN_GREY, bottom=THIN_GREY)

COLUMN_WIDTHS = {
    "会社ID": 13,
    "会社名": 30,
    "地域": 14,
    "都道府県": 14,
    "本社所在地": 32,
    "営業エリア": 38,
    "戸建て取扱": 14,
    "収益不動産取扱": 18,
    "その他取扱物件": 42,
    "問い合わせフォーム": 20,
    "公式URL": 36,
    "問い合わせURL": 42,
    "サービスURL": 42,
    "電話番号": 18,
    "特徴・強み": 46,
    "根拠URL": 52,
    "確認日": 14,
    "確認状態": 14,
    "優先度": 12,
    "備考": 44,
}


def _table_name(title: str) -> str:
    digest = hashlib.sha1(title.encode("utf-8"), usedforsecurity=False).hexdigest()[:12]
    return f"T_{digest}"


def _write_sheet(ws, rows: list[dict[str, str]], title: str) -> None:
    ws.title = title[:31]
    ws.append(REQUIRED_COLUMNS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[1].height = 34

    for row in rows:
        ws.append([row.get(column, "") for column in REQUIRED_COLUMNS])

    for row_cells in ws.iter_rows(min_row=2):
        ws.row_dimensions[row_cells[0].row].height = 42
        for cell in row_cells:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = BORDER
        for column in URL_COLUMNS:
            index = REQUIRED_COLUMNS.index(column) + 1
            cell = row_cells[index - 1]
            value = str(cell.value or "")
            first_url = next(
                (part.strip() for part in value.split("|") if part.strip().startswith("https://")),
                None,
            )
            if first_url:
                cell.hyperlink = first_url
                cell.font = LINK_FONT

    for index, column in enumerate(REQUIRED_COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(index)].width = COLUMN_WIDTHS.get(column, 18)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.sheet_view.showGridLines = False
    if ws.max_row >= 2:
        table = Table(displayName=_table_name(title), ref=ws.dimensions)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)


def _write_dictionary(ws) -> None:
    ws.title = "データ辞書"
    ws.append(["列名", "説明"])
    descriptions = {
        "会社ID": "重複しない管理ID",
        "会社名": "法人・ブランドの正式名称",
        "地域": "地域別シート分類",
        "都道府県": "本社所在地または主担当拠点",
        "本社所在地": "公式情報で確認した所在地",
        "営業エリア": "主な対応エリア",
        "戸建て取扱": "あり / なし / 要確認",
        "収益不動産取扱": "あり / なし / 要確認",
        "その他取扱物件": "マンション、土地、事業用、賃貸管理等",
        "問い合わせフォーム": "あり / なし / 要確認",
        "公式URL": "会社・ブランド公式サイト",
        "問い合わせURL": "問い合わせフォームまたは問い合わせ案内",
        "サービスURL": "取扱物件の根拠となるサービスページ",
        "電話番号": "公開されている代表・相談窓口",
        "特徴・強み": "公式サイトから確認できる特徴",
        "根拠URL": "確認に使用した公式ページ。複数は | 区切り",
        "確認日": "最終確認日 YYYY-MM-DD",
        "確認状態": "確認済み / 一部確認 / 要確認",
        "優先度": "A / B / C",
        "備考": "注意点、追加調査事項",
    }
    for column in REQUIRED_COLUMNS:
        ws.append([column, descriptions.get(column, "")])
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 80
    ws.freeze_panes = "A2"


def _write_summary(rows: list[dict[str, str]], output: Path) -> None:
    region_counts = Counter(row["地域"] for row in rows)
    detached = sum(row["戸建て取扱"] == "あり" for row in rows)
    income = sum(row["収益不動産取扱"] == "あり" for row in rows)
    forms = sum(row["問い合わせフォーム"] == "あり" for row in rows)
    lines = [
        "# 不動産取扱業者データベース 集計",
        "",
        f"- 総登録数: {len(rows)}社",
        f"- 戸建て取扱確認済み: {detached}社",
        f"- 収益不動産取扱確認済み: {income}社",
        f"- 問い合わせフォーム確認済み: {forms}社",
        "",
        "## 地域別",
        "",
    ]
    for region in REGION_ORDER:
        if region_counts[region]:
            lines.append(f"- {region}: {region_counts[region]}社")
    output.write_text("\n".join(lines) + "\n", encoding="utf-8")


def build(input_path: Path, output_dir: Path) -> Path:
    rows = validate_file(input_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb.create_sheet(), rows, "全社一覧")
    for region in REGION_ORDER:
        region_rows = [row for row in rows if row["地域"] == region]
        if region_rows:
            _write_sheet(wb.create_sheet(), region_rows, region)
    _write_dictionary(wb.create_sheet())

    xlsx_path = output_dir / "real_estate_brokers.xlsx"
    wb.save(xlsx_path)

    for filename in ["real_estate_brokers.csv", "notion_import.csv"]:
        with (output_dir / filename).open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=REQUIRED_COLUMNS)
            writer.writeheader()
            writer.writerows(rows)

    _write_summary(rows, output_dir / "summary.md")
    return xlsx_path


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT)
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR)
    args = parser.parse_args()
    path = build(args.input, args.output_dir)
    print(path)


if __name__ == "__main__":
    main()
