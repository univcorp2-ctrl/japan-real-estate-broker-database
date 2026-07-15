from pathlib import Path

from openpyxl import load_workbook

from real_estate_db.build_excel import build


def test_build_creates_expected_files(tmp_path: Path) -> None:
    xlsx = build(Path("data/real_estate_brokers.csv"), tmp_path)
    assert xlsx.exists()
    assert (tmp_path / "real_estate_brokers.csv").exists()
    assert (tmp_path / "notion_import.csv").exists()
    assert (tmp_path / "summary.md").exists()

    workbook = load_workbook(xlsx)
    assert "全社一覧" in workbook.sheetnames
    assert "関東" in workbook.sheetnames
    assert "九州・沖縄" in workbook.sheetnames
    sheet = workbook["全社一覧"]
    assert sheet.freeze_panes == "A2"
    assert sheet.max_row >= 21
    assert len(sheet.tables) == 1
